from pathlib import Path

import pdfplumber
from docx import Document
from openpyxl import load_workbook


MAX_PDF_PAGES = 500
MAX_PDF_CHARS = 2_000_000


def extract_text_from_pdf(file_path: str | Path) -> dict:
    try:
        path = Path(file_path)
        pages = []
        with pdfplumber.open(path) as pdf:
            page_count = len(pdf.pages)
            if page_count > MAX_PDF_PAGES:
                return {
                    "text": "",
                    "page_count": page_count,
                    "is_image_only": False,
                    "error": f"PDF has {page_count} pages which exceeds the {MAX_PDF_PAGES}-page limit.",
                }
            total_chars = 0
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                pages.append(page_text)
                total_chars += len(page_text)
                if total_chars > MAX_PDF_CHARS:
                    pages.append("[TRUNCATED: document exceeds character limit]")
                    break
        text = "\n".join(pages)
        is_image_only = len(text.strip()) < 100
        return {"text": text, "page_count": page_count, "is_image_only": is_image_only, "error": None}
    except Exception as e:
        return {"text": "", "page_count": 0, "is_image_only": False, "error": str(e)}


def extract_text_from_docx(file_path: str | Path) -> dict:
    try:
        path = Path(file_path)
        doc = Document(path)
        text = "\n".join(p.text for p in doc.paragraphs)
        return {"text": text, "page_count": None, "is_image_only": False, "error": None}
    except Exception as e:
        return {"text": "", "page_count": None, "is_image_only": False, "error": str(e)}


def extract_text_from_xlsx(file_path: str | Path) -> dict:
    try:
        path = Path(file_path)
        wb = load_workbook(path, read_only=True)
        sheet_names = wb.sheetnames
        lines = []
        for name in sheet_names:
            lines.append(f"## Sheet: {name}")
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append("\t".join(cells))
        wb.close()
        return {"text": "\n".join(lines), "sheet_names": sheet_names, "error": None}
    except Exception as e:
        return {"text": "", "sheet_names": [], "error": str(e)}


_CANNOT_AUTO_PROCESS = {".dwg", ".vsdx", ".rvt", ".nwc", ".ifc", ".msg", ".eml"}
_SUPPORTED = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".xlsm"}


def extract_text(file_path: str | Path) -> dict:
    ext = Path(file_path).suffix.lower()

    if ext in _CANNOT_AUTO_PROCESS:
        return {"text": "", "error": "cannot_auto_process", "can_auto_process": False}

    if ext not in _SUPPORTED:
        return {"text": "", "error": "unsupported_format", "can_auto_process": False}

    if ext == ".pdf":
        result = extract_text_from_pdf(file_path)
    elif ext in {".docx", ".doc"}:
        result = extract_text_from_docx(file_path)
    else:
        result = extract_text_from_xlsx(file_path)

    result["can_auto_process"] = True
    return result
