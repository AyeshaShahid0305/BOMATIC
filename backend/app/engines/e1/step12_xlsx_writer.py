from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_STATUS_FILLS = {
    "Compliant":          PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "Partial":            PatternFill(fill_type="solid", fgColor="FFEB9C"),
    "Non-Compliant":      PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "Alternative Offered": PatternFill(fill_type="solid", fgColor="DDEBF7"),
}

_MATRIX_COLUMNS: list[tuple[str, int]] = [
    ("Req #",                  8),
    ("RFP Requirement Text",  45),
    ("Classification",        15),
    ("Framework",             12),
    ("Control ID",            12),
    ("Control Name",          25),
    ("Status",                18),
    ("TP Section",            25),
    ("Notes",                 30),
    ("Gap Type",              18),
]

_MATRIX_ROW_KEYS = [
    "req_id", "req_text", "classification", "framework",
    "control_id", "control_name", "status", "tp_section", "notes", "gap_type",
]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _write_matrix_sheet(ws, matrix_rows: list[dict]) -> None:
    ws.title = "Compliance Matrix"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Header
    for col_idx, (header, width) in enumerate(_MATRIX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row in enumerate(matrix_rows, start=2):
        for col_idx, key in enumerate(_MATRIX_ROW_KEYS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Conditional status colour (column 7)
        status = row.get("status", "")
        fill = _STATUS_FILLS.get(status)
        if fill:
            ws.cell(row=row_idx, column=7).fill = fill


def _write_gap_sheet(ws, gaps: dict) -> None:
    ws.title = "Gap Analysis"

    coverage_gaps: list[dict] = gaps.get("coverage_gaps", [])
    orphan_reqs: list[dict] = gaps.get("orphan_requirements", [])

    header_font = Font(bold=True)

    # --- Table 1: Coverage Gaps ---
    ws.cell(row=1, column=1, value="Coverage Gaps (framework controls with no matching requirement)").font = Font(bold=True, size=12)

    gap_headers = ["Framework", "Control ID", "Control Name"]
    for col_idx, h in enumerate(gap_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")

    for row_idx, gap in enumerate(coverage_gaps, start=3):
        ws.cell(row=row_idx, column=1, value=gap.get("framework", ""))
        ws.cell(row=row_idx, column=2, value=gap.get("control_id", ""))
        ws.cell(row=row_idx, column=3, value=gap.get("control_name", ""))

    # --- Table 2: Orphan Requirements ---
    orphan_start = len(coverage_gaps) + 5
    ws.cell(row=orphan_start, column=1, value="Orphan Requirements (RFP requirements with no matching control)").font = Font(bold=True, size=12)

    orphan_headers = ["Req #", "Requirement Text", "Classification"]
    for col_idx, h in enumerate(orphan_headers, start=1):
        cell = ws.cell(row=orphan_start + 1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL

    for row_idx, req in enumerate(orphan_reqs, start=orphan_start + 2):
        ws.cell(row=row_idx, column=1, value=req.get("req_id", ""))
        ws.cell(row=row_idx, column=2, value=req.get("req_text", ""))
        ws.cell(row=row_idx, column=3, value=req.get("classification", ""))

    # Column widths
    for col, width in [(1, 18), (2, 50), (3, 18)]:
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_stats_sheet(ws, stats: dict) -> None:
    ws.title = "Stats"

    ws.cell(row=1, column=1, value="Metric").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=2, value="Value").font = Font(bold=True, color="FFFFFF")
    for col in [1, 2]:
        ws.cell(row=1, column=col).fill = _HEADER_FILL

    stat_labels = [
        ("total_reqs",      "Total Requirements"),
        ("compliant",       "Compliant"),
        ("partial",         "Partial"),
        ("non_compliant",   "Non-Compliant"),
        ("alternative",     "Alternative Offered"),
        ("orphans",         "Orphan Requirements"),
        ("coverage_gaps",   "Coverage Gaps"),
    ]

    for row_idx, (key, label) in enumerate(stat_labels, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=stats.get(key, 0))

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_compliance_matrix_xlsx(
    matrix_rows: list[dict],
    gaps: dict,
    stats: dict,
    opportunity_id: str,
    output_dir: Path,
) -> Path:
    """
    Write the compliance matrix to an .xlsx file and return the full path.
    File is named: e1_{opportunity_id}_compliance_matrix_v1.xlsx
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"e1_{opportunity_id}_compliance_matrix_v1.xlsx"

    wb = Workbook()
    ws_matrix = wb.active
    ws_gap = wb.create_sheet()
    ws_stats = wb.create_sheet()

    _write_matrix_sheet(ws_matrix, matrix_rows)
    _write_gap_sheet(ws_gap, gaps)
    _write_stats_sheet(ws_stats, stats)

    wb.save(output_path)
    return output_path
