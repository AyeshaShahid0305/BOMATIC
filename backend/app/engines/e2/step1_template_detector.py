import re
from pathlib import Path

import openpyxl

from .models import BoQDetectionResult

_CISCO_SKU_RE = re.compile(r"^[A-Z][A-Z0-9]{1,}-[A-Z0-9]")

FORMAT_1_CCW = "FORMAT_1_CCW"
UNKNOWN = "UNKNOWN"
FORMAT_2_ARAMCO = "FORMAT_2_ARAMCO"
FORMAT_3_NTT = "FORMAT_3_NTT"


def detect_template(file_path: Path) -> BoQDetectionResult:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames

        # 1. Sheet name match
        for name in sheet_names:
            if name.lower() == "main boq":
                ws = wb[name]
                _get_row_values(ws, 1)
                return BoQDetectionResult(
                    format_type=FORMAT_1_CCW,
                    confidence=0.9,
                    sheet_name=name,
                    header_row_index=0,
                )

        # 2. Header match — all required headers present
        for name in sheet_names:
            ws = wb[name]
            header_row = _get_row_values(ws, 1)
            if (
                "Part Number" in header_row
                and "Description" in header_row
                and "Qty" in header_row
                and any(h.startswith("Unit Price") for h in header_row)
                and any(h.startswith("Total Price") for h in header_row)
            ):
                return BoQDetectionResult(
                    format_type=FORMAT_1_CCW,
                    confidence=0.95,
                    sheet_name=name,
                    header_row_index=0,
                )

        # 3. SKU regex fallback — first data cell of each sheet
        for name in sheet_names:
            ws = wb[name]
            first_data_cell = _get_cell_value(ws, 2, 1)
            if first_data_cell and _CISCO_SKU_RE.match(str(first_data_cell)):
                return BoQDetectionResult(
                    format_type=FORMAT_1_CCW,
                    confidence=0.7,
                    sheet_name=name,
                    header_row_index=0,
                )

        # 4. Aramco format — sheet name or header pattern
        ARAMCO_SHEET_NAMES = {'boq', 'bill of quantities', 'boq sheet', 'price schedule', 'price list'}
        ARAMCO_HEADERS = {'item no', 'item description', 'unit rate', 'quantity', 'amount'}
        for name in sheet_names:
            if name.lower() in ARAMCO_SHEET_NAMES:
                return BoQDetectionResult(
                    format_type=FORMAT_2_ARAMCO,
                    confidence=0.9,
                    sheet_name=name,
                    header_row_index=0,
                )
        for name in sheet_names:
            ws = wb[name]
            header_row = [h.lower() for h in _get_row_values(ws, 1)]
            matched = sum(1 for h in ARAMCO_HEADERS if any(h in cell for cell in header_row))
            if matched >= 3:
                return BoQDetectionResult(
                    format_type=FORMAT_2_ARAMCO,
                    confidence=0.7 + 0.05 * matched,
                    sheet_name=name,
                    header_row_index=0,
                )
        # 5. NTT format — sheet name or header pattern
        NTT_SHEET_NAMES = {'bom', 'bill of materials', 'pricing', 'price sheet', 'ntt boq'}
        NTT_HEADERS = {'part no', 'product description', 'list price', 'net price', 'quantity', 'item'}
        for name in sheet_names:
            if name.lower() in NTT_SHEET_NAMES:
                return BoQDetectionResult(
                    format_type=FORMAT_3_NTT,
                    confidence=0.9,
                    sheet_name=name,
                    header_row_index=0,
                )
        for name in sheet_names:
            ws = wb[name]
            header_row = [h.lower() for h in _get_row_values(ws, 1)]
            matched = sum(1 for h in NTT_HEADERS if any(h in cell for cell in header_row))
            if matched >= 3:
                return BoQDetectionResult(
                    format_type=FORMAT_3_NTT,
                    confidence=0.7 + 0.05 * matched,
                    sheet_name=name,
                    header_row_index=0,
                )
        return BoQDetectionResult(
            format_type=UNKNOWN,
            confidence=0.5,
            sheet_name=sheet_names[0] if sheet_names else "",
            header_row_index=0,
        )
    finally:
        wb.close()


def _get_row_values(ws, row_index: int) -> list[str]:
    row = next(
        (r for i, r in enumerate(ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True), start=1) if i == 1),
        None,
    )
    if row is None:
        return []
    return [str(v).strip() for v in row if v is not None]


def _get_cell_value(ws, row: int, col: int):
    for i, r in enumerate(ws.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col, values_only=True), start=1):
        if i == 1:
            return r[0]
    return None


if __name__ == "__main__":
    fixture = Path(__file__).parents[3] / "storage" / "e1_test_fixtures" / "BOQ_4203193153.xlsx"
    result = detect_template(fixture)
    print(result)
