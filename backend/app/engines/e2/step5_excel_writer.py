from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import BoQDetectionResult, PricingSummary
from .step1_template_detector import UNKNOWN
from .step2_boq_parser import COLUMN_MAPPINGS

_OUTPUT_DIR = Path(__file__).parent / "output"
_SI_DISCOUNT = 0.15

_HEADER_COLS = [
    "Part Number", "Description", "Qty", "Unit Price",
    "Discount %", "Line Total", "Match Method", "Match Score",
]

_BOLD = Font(bold=True)
_BOLD_WHITE = Font(bold=True, color="FFFFFF")
_RIGHT = Alignment(horizontal="right")
_RED_FILL = PatternFill(fill_type="solid", fgColor="FF0000")


def write_output(
    summary: PricingSummary,
    template_path: Path,
    detection: BoQDetectionResult,
    cost_stack_result: dict | None = None,
) -> str:
    _OUTPUT_DIR.mkdir(exist_ok=True)
    has_cost_stack = bool(cost_stack_result and cost_stack_result.get("line_items"))
    currency = (cost_stack_result or {}).get("summary", {}).get("currency", "SAR") if has_cost_stack else "SAR"

    header_cols = _HEADER_COLS.copy()
    if has_cost_stack:
        header_cols += [
            f"List ({currency})",
            f"After Discount ({currency})",
            f"After Margin ({currency})",
            f"Overhead ({currency})",
            f"Cost+OH ({currency})",
            f"Sell Price ({currency})",
            f"Extended Sell ({currency})",
            f"VAT ({currency})",
            f"Total incl. VAT ({currency})",
        ]

    stem = Path(template_path).stem
    output_path = _OUTPUT_DIR / f"{stem}_BOMATIC_filled.xlsx"

    wb = openpyxl.load_workbook(template_path)
    ws = wb[detection.sheet_name]

    # detection.header_row_index is 0-based; openpyxl rows are 1-based
    header_row = detection.header_row_index + 1

    if detection.format_type != UNKNOWN and detection.format_type in COLUMN_MAPPINGS:
        _write_mapped_client_rows(ws, summary, detection, header_row)
        wb.save(output_path)
        return str(output_path)

    # Write our column headers over the template's original header row
    for col, title in enumerate(header_cols, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = _BOLD

    # Find the first empty data row after the header
    row = _first_empty_row(ws, header_row + 1)

    cs_by_sku: dict[str, dict] = {}
    if has_cost_stack:
        for item in cost_stack_result["line_items"]:
            cs_by_sku[item["sku"]] = item

    # --- Matched items ---
    for m in summary.matched_items:
        qty = m.rfp_item.quantity if m.rfp_item.quantity is not None else 1.0
        line_total = round(qty * m.unit_price * (1 - _SI_DISCOUNT), 2)
        for col, val in enumerate(
            [m.sku, m.rfp_item.description, qty, m.unit_price,
             f"{_SI_DISCOUNT:.0%}", line_total, m.match_method, m.match_score],
            start=1,
        ):
            ws.cell(row=row, column=col, value=val)

        # Cost stack columns
        if has_cost_stack and m.sku in cs_by_sku:
            cs = cs_by_sku[m.sku]
            col = 9
            for val in [
                cs["cs001_line_list_local"],
                cs["cs002_after_vendor_discount"],
                cs["cs003_after_inhouse_margin"],
                cs["cs004_overhead"],
                cs["cs005_cost_with_overhead"],
                cs["cs006_selling_price"],
                cs["cs007_extended_sell"],
                cs["cs009_vat"],
                cs["line_total_with_vat"],
            ]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = _RIGHT
                col += 1

        row += 1

    row += 1  # blank separator

    # --- Unmatched items (NEEDS REVIEW) ---
    for m in summary.unmatched_items:
        review_cell = ws.cell(row=row, column=1, value="NEEDS REVIEW")
        review_cell.fill = _RED_FILL
        review_cell.font = _BOLD_WHITE
        ws.cell(row=row, column=2, value=m.rfp_item.description)
        if m.rfp_item.quantity is not None:
            ws.cell(row=row, column=3, value=m.rfp_item.quantity)
        ws.cell(row=row, column=7, value="unmatched")
        row += 1

    row += 1  # blank separator before summary

    # --- Summary block (columns G–H, bold + right-aligned) ---
    for label, value in [
        ("Subtotal", summary.subtotal),
        (f"Discount Amount ({_SI_DISCOUNT:.0%})", summary.discount_amount),
        ("Total", summary.total),
    ]:
        label_cell = ws.cell(row=row, column=7, value=label)
        label_cell.font = _BOLD
        label_cell.alignment = _RIGHT

        value_cell = ws.cell(row=row, column=8, value=f"{summary.currency} {value:,.2f}")
        value_cell.font = _BOLD
        value_cell.alignment = _RIGHT
        row += 1

    if has_cost_stack:
        cs_summary = cost_stack_result["summary"]
        row += 1
        ws.cell(row=row, column=7, value=f" Cost Stack Summary ({cs_summary['currency']}) ").font = _BOLD
        row += 1
        for label, value in [
            (f"Total List ({cs_summary['currency']})", cs_summary["total_list_local"]),
            ("Total Cost (with OH)", cs_summary["total_cost"]),
            ("Total Sell Price", cs_summary["total_sell"]),
            (f"VAT ({cs_summary['vat_rate']:.0%})", cs_summary["total_vat"]),
            ("Total incl. VAT", cs_summary["total_with_vat"]),
            (f"STCS Sale ({cs_summary['vat_country']})", cs_summary["cs008_stcs_sale_total"]),
        ]:
            lc = ws.cell(row=row, column=7, value=label)
            lc.font = _BOLD
            lc.alignment = _RIGHT
            vc = ws.cell(row=row, column=8, value=f"{cs_summary['currency']} {value:,.2f}")
            vc.font = _BOLD
            vc.alignment = _RIGHT
            row += 1

    wb.save(output_path)
    return str(output_path)


def _write_mapped_client_rows(
    ws,
    summary: PricingSummary,
    detection: BoQDetectionResult,
    header_row: int,
) -> None:
    mapping = COLUMN_MAPPINGS[detection.format_type]
    data_rows = [
        row_index
        for row_index in range(header_row + 1, ws.max_row + 1)
        if any(
            ws.cell(row=row_index, column=column_index + 1).value is not None
            for column_index in mapping.values()
        )
    ]
    rows_by_sku = {
        str(ws.cell(row=row_index, column=mapping["part_number"] + 1).value).strip().lower(): row_index
        for row_index in data_rows
        if ws.cell(row=row_index, column=mapping["part_number"] + 1).value is not None
    }
    used_rows: set[int] = set()

    for match in summary.matched_items:
        row_index = rows_by_sku.get(match.sku.strip().lower())
        if row_index in used_rows:
            row_index = None
        if row_index is None:
            row_index = next((row for row in data_rows if row not in used_rows), None)
        if row_index is None:
            break

        used_rows.add(row_index)
        qty = match.rfp_item.quantity if match.rfp_item.quantity is not None else 1.0
        line_total = round(qty * match.unit_price * (1 - _SI_DISCOUNT), 2)
        values = {
            "part_number": match.sku,
            "description": match.rfp_item.description,
            "qty": qty,
            "unit_price": match.unit_price,
            "total_price": line_total,
        }
        for field, value in values.items():
            ws.cell(row=row_index, column=mapping[field] + 1, value=value)


def _first_empty_row(ws, from_row: int) -> int:
    """Return the first 1-based row index (>= from_row) where all cells are empty."""
    for idx, row in enumerate(ws.iter_rows(min_row=from_row, values_only=True), start=from_row):
        if all(v is None for v in row):
            return idx
    return ws.max_row + 1


def write_distributor_export(
    summary: PricingSummary,
    template_path: Path,
) -> Path:
    """
    Write a clean distributor-facing Excel file containing only matched items.

    Columns: SKU | Description | Quantity | Unit Price (USD) | Total Price (USD)
    Includes a bold totals row. Unmatched items are excluded.
    Returns the path to the generated file.
    """
    _OUTPUT_DIR.mkdir(exist_ok=True)

    stem = Path(template_path).stem
    output_path = _OUTPUT_DIR / f"{stem}_Distributor_Export.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Distributor Order"

    # Header row
    headers = ["SKU", "Description", "Quantity", "Unit Price (USD)", "Total Price (USD)"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _BOLD

    # Data rows - matched items only
    row = 2
    total_qty = 0.0
    total_price = 0.0

    for m in summary.matched_items:
        qty = m.rfp_item.quantity if m.rfp_item.quantity is not None else 1.0
        line_total = round(qty * m.unit_price, 2)
        total_qty += qty
        total_price += line_total

        ws.cell(row=row, column=1, value=m.sku)
        ws.cell(row=row, column=2, value=m.rfp_item.description)
        ws.cell(row=row, column=3, value=qty)
        ws.cell(row=row, column=4, value=round(m.unit_price, 2))
        ws.cell(row=row, column=5, value=line_total)
        row += 1

    # Totals row
    totals_cells = [
        (row, 1, ""),
        (row, 2, "TOTAL"),
        (row, 3, total_qty),
        (row, 4, ""),
        (row, 5, round(total_price, 2)),
    ]
    for r, c, v in totals_cells:
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = _BOLD

    # Auto-fit column widths (best-effort)
    col_widths = [16, 45, 10, 18, 18]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    from app.engines.e2.models import CatalogMatch, RFPLineItem
    from app.engines.e2.step4_gap_analyzer import analyze_gaps

    matches = [
        CatalogMatch(
            rfp_item=RFPLineItem(description="Cisco ASA 5516-X firewall", quantity=2, unit="units", category="security", raw_text="", confidence=0.9),
            sku="ASA5516-FPWR-K9", product_name="Cisco ASA 5516-X with FirePOWER Services",
            vendor="Cisco", unit_price=4995.00, match_score=1.0, match_method="exact",
        ),
        CatalogMatch(
            rfp_item=RFPLineItem(description="48-port PoE+ switch", quantity=10, unit="units", category="network", raw_text="", confidence=0.85),
            sku="C9300-48P-E", product_name="Cisco Catalyst 9300 48-Port PoE+ Switch",
            vendor="Cisco", unit_price=7200.00, match_score=0.55, match_method="fuzzy",
        ),
        CatalogMatch(
            rfp_item=RFPLineItem(description="unmanaged desktop hub", quantity=1, unit="units", category="hardware", raw_text="", confidence=0.6),
            sku="", product_name="", vendor="", unit_price=0.0, match_score=0.0, match_method="unmatched",
        ),
    ]

    summary = analyze_gaps(matches)

    fixture = Path(__file__).parents[3] / "storage" / "e1_test_fixtures" / "BOQ_4203193153.xlsx"
    from app.engines.e2.step1_template_detector import detect_template
    detection = detect_template(fixture)

    out = write_output(summary, fixture, detection)
    print(f"Written: {out}")
