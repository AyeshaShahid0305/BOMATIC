import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import pdfplumber

from app.schemas.pipeline import E4BaselineArtifact, E4BaselineRequirement

_INSUFFICIENT_ANSWERS = {
    "n/a",
    "na",
    "none",
    "tbd",
    "unknown",
    "to be confirmed",
    "not available",
}


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _is_insufficient(answer: str, expected_type: str) -> bool:
    normalized = answer.strip().lower()
    if not normalized:
        return False
    if normalized in _INSUFFICIENT_ANSWERS or len(normalized) < 2:
        return True
    if expected_type == "number" and not re.search(r"\d", normalized):
        return True
    if expected_type == "yes_no" and normalized not in {"yes", "no", "y", "n"}:
        return True
    return False


def _rows_from_xlsx(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["RFI Questionnaire"] if "RFI Questionnaire" in workbook.sheetnames else workbook.active
        raw_rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not raw_rows:
        return []
    headers = {_text(value).lower(): index for index, value in enumerate(raw_rows[0])}

    def cell(row, *names):
        for name in names:
            index = headers.get(name)
            if index is not None and index < len(row):
                return _text(row[index])
        return ""

    return [
        {
            "question_id": cell(row, "rfi id", "question id", "id"),
            "category": cell(row, "category"),
            "priority": cell(row, "priority"),
            "question": cell(row, "question"),
            "expected_answer_type": cell(row, "answer type", "expected answer type"),
            "answer": cell(row, "response", "answer"),
        }
        for row in raw_rows[1:]
        if any(_text(value) for value in row)
    ]


def _rows_from_pdf(path: Path) -> list[dict]:
    rows: list[dict] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table:
                    continue
                headers = {_text(value).lower(): index for index, value in enumerate(table[0])}
                for row in table[1:]:
                    def cell(*names):
                        for name in names:
                            index = headers.get(name)
                            if index is not None and index < len(row):
                                return _text(row[index])
                        return ""

                    rows.append(
                        {
                            "question_id": cell("rfi id", "question id", "id"),
                            "category": cell("category"),
                            "priority": cell("priority"),
                            "question": cell("question"),
                            "expected_answer_type": cell("answer type", "expected answer type"),
                            "answer": cell("response", "answer"),
                        }
                    )

            text = page.extract_text() or ""
            for line in text.splitlines():
                match = re.match(r"^\s*(RFI-\d+)\s*[|:\-]\s*(.+)$", line, re.IGNORECASE)
                if match and not any(row["question_id"].lower() == match.group(1).lower() for row in rows):
                    rows.append({"question_id": match.group(1), "answer": match.group(2)})
    return rows


def parse_rfi_response(
    path: Path,
    questions: list[dict],
    source_filename: str,
) -> E4BaselineArtifact:
    if path.suffix.lower() == ".xlsx":
        response_rows = _rows_from_xlsx(path)
    elif path.suffix.lower() == ".pdf":
        response_rows = _rows_from_pdf(path)
    else:
        raise ValueError("RFI response must be an XLSX or PDF file.")

    source_questions = questions or response_rows
    rows_by_id = {
        row.get("question_id", "").strip().lower(): row
        for row in response_rows
        if row.get("question_id", "").strip()
    }
    rows_by_category: dict[str, list[dict]] = defaultdict(list)
    for row in response_rows:
        rows_by_category[row.get("category", "").strip().lower()].append(row)
    category_offsets: dict[str, int] = defaultdict(int)

    requirements: list[E4BaselineRequirement] = []
    for index, question in enumerate(source_questions):
        question_id = _text(question.get("id") or question.get("question_id") or f"RFI-{index + 1:03d}")
        category = _text(question.get("category") or "General")
        row = rows_by_id.get(question_id.lower())
        if row is None:
            category_key = category.lower()
            offset = category_offsets[category_key]
            candidates = rows_by_category.get(category_key, [])
            row = candidates[offset] if offset < len(candidates) else None
            category_offsets[category_key] += 1
        if row is None and index < len(response_rows):
            row = response_rows[index]

        answer = _text((row or {}).get("answer"))
        expected_type = _text(
            question.get("expected_answer_type")
            or (row or {}).get("expected_answer_type")
            or "text"
        )
        if not answer:
            status = "missing"
            gap_reason = "No answer provided."
        elif _is_insufficient(answer, expected_type):
            status = "insufficient"
            gap_reason = "Answer does not satisfy the expected answer type or level of detail."
        else:
            status = "answered"
            gap_reason = None

        requirements.append(
            E4BaselineRequirement(
                question_id=question_id,
                category=category,
                question=_text(question.get("question") or (row or {}).get("question")),
                answer=answer,
                priority=_text(question.get("priority") or (row or {}).get("priority") or "must_have"),
                expected_answer_type=expected_type,
                status=status,
                gap_reason=gap_reason,
            )
        )

    answered_count = sum(item.status == "answered" for item in requirements)
    return E4BaselineArtifact(
        source_filename=source_filename,
        requirements=requirements,
        answered_count=answered_count,
        gap_count=len(requirements) - answered_count,
    )
