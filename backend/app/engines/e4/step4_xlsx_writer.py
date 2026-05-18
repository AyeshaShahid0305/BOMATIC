import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import RFIQuestionnaire

_OUTPUT_DIR = Path(__file__).parent / "output"

_COL_WIDTHS = [10, 15, 12, 60, 40, 15, 40]
_HEADERS = ["RFI ID", "Category", "Priority", "Question", "Rationale", "Answer Type", "Response"]

_DARK_BLUE_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFACD")
_WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")


def write_questionnaire(questionnaire: RFIQuestionnaire) -> str:
    _OUTPUT_DIR.mkdir(exist_ok=True)

    safe_name = re.sub(r'[\\/*?:"<>|]', "_", questionnaire.project_name or "RFI")
    output_path = _OUTPUT_DIR / f"{safe_name}_RFI.xlsx"

    wb = openpyxl.Workbook()

    _write_questionnaire_sheet(wb.active, questionnaire)
    wb.active.title = "RFI Questionnaire"

    summary_ws = wb.create_sheet("Summary")
    _write_summary_sheet(summary_ws, questionnaire)

    wb.save(output_path)
    return str(output_path)


def _write_questionnaire_sheet(ws, questionnaire: RFIQuestionnaire) -> None:
    # Headers
    for col, (title, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _DARK_BLUE_FILL
        cell.alignment = _WRAP_CENTER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, q in enumerate(questionnaire.questions, start=2):
        row_fill = _YELLOW_FILL if q.priority == "must_have" else _WHITE_FILL
        values = [q.id, q.category, q.priority, q.question, q.rationale, q.expected_answer_type, ""]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = _WRAP

    ws.freeze_panes = "A2"


def _write_summary_sheet(ws, questionnaire: RFIQuestionnaire) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    meta_rows = [
        ("Project Name", questionnaire.project_name),
        ("Generated From", questionnaire.generated_from),
        ("Date", str(date.today())),
        ("Total Questions", questionnaire.total_questions),
    ]

    for row_idx, (label, value) in enumerate(meta_rows, start=1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = _BOLD
        ws.cell(row=row_idx, column=2, value=value)

    # Category breakdown table
    ws.cell(row=6, column=1, value="Category").font = _BOLD
    ws.cell(row=6, column=2, value="Question Count").font = _BOLD

    counts: dict[str, int] = {}
    for q in questionnaire.questions:
        counts[q.category] = counts.get(q.category, 0) + 1

    for row_idx, category in enumerate(questionnaire.categories, start=7):
        ws.cell(row=row_idx, column=1, value=category)
        ws.cell(row=row_idx, column=2, value=counts.get(category, 0))
