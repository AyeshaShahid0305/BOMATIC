"""
Run from backend/ directory:
    .venv/Scripts/python.exe tests/create_fixtures.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

FIXTURES_DIR = Path(__file__).parent / "fixtures"

HEADERS = ["Part Number", "Description", "Qty", "Unit Price", "Total Price"]

ROWS = [
    ("ASA5516-FPWR-K9",   "Cisco ASA 5516-X with FirePOWER Services, 8GE, AC, 3DES/AES",          10, None, None),
    ("L-ASA5516-TAMC=",   "Cisco ASA5516 Threat, AMP and URL 3-Year Subscription",                 10, None, None),
    ("C9300-48P-E",       "Cisco Catalyst 9300 48-Port PoE+ Network Essentials",                   20, None, None),
    ("C9300-NM-4G",       "Cisco Catalyst 9300 4x1G Network Module",                               20, None, None),
    ("ISR4331/K9",        "Cisco 4331 Integrated Services Router",                                  5,  None, None),
    ("FL-4330-HSEC-K9=",  "Cisco 4330 U.S. Export Restriction Compliance licence",                  5,  None, None),
    ("AIR-AP2802I-E-K9",  "Cisco Aironet 2802I 802.11ac Wave 2 AP, Internal Antenna",              50, None, None),
    ("DN2-HW-APL",        "Cisco DNA Centre 44-Core Appliance",                                     1,  None, None),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=11)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COL_WIDTHS = [22, 60, 10, 14, 14]


def build_boq():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of Quantities"
    ws.row_dimensions[1].height = 28

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = width

    for row_num, row_data in enumerate(ROWS, start=2):
        ws.row_dimensions[row_num].height = 36
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = LEFT if col_idx == 2 else CENTER

    out = FIXTURES_DIR / "sample_boq.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    path = build_boq()
    print(f"Created: {path}  ({path.stat().st_size:,} bytes)")
