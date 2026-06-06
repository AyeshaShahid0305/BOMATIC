import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

import app.engines.e2.step5_excel_writer as writer
from app.engines.e2.models import CatalogMatch, PricingSummary, RFPLineItem
from app.engines.e2.step1_template_detector import detect_template


def test_write_output_preserves_ccw_headers_and_updates_mapped_row(tmp_path, monkeypatch):
    template_path = tmp_path / "ccw_template.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Main BOQ"
    original_headers = [
        "Part Number",
        "Description",
        "Qty",
        "Unit Price (USD)",
        "Total Price (USD)",
    ]
    worksheet.append(original_headers)
    worksheet.append(["C9300-48P-E", "Original description", 1, 0, 0])
    workbook.save(template_path)

    rfp_item = RFPLineItem(
        description="Cisco Catalyst 9300 48-Port PoE+ Switch",
        quantity=2,
        unit="units",
        category="network",
        raw_text="",
        confidence=1.0,
    )
    match = CatalogMatch(
        rfp_item=rfp_item,
        sku="C9300-48P-E",
        product_name="Cisco Catalyst 9300 48-Port PoE+ Switch",
        vendor="Cisco",
        unit_price=7200,
        match_score=1.0,
        match_method="exact",
    )
    summary = PricingSummary(
        matched_items=[match],
        subtotal=14400,
        discount_amount=2160,
        total=12240,
    )
    detection = detect_template(template_path)
    monkeypatch.setattr(writer, "_OUTPUT_DIR", tmp_path / "output")

    output_path = writer.write_output(summary, template_path, detection)

    output_workbook = openpyxl.load_workbook(output_path, data_only=True)
    output_sheet = output_workbook["Main BOQ"]
    assert [output_sheet.cell(row=1, column=column).value for column in range(1, 6)] == original_headers
    assert output_sheet["A2"].value == "C9300-48P-E"
    assert output_sheet["B2"].value == rfp_item.description
    assert output_sheet["C2"].value == 2
    assert output_sheet["D2"].value == 7200
    assert output_sheet["E2"].value == 12240
    assert output_sheet.max_row == 2
    output_workbook.close()
