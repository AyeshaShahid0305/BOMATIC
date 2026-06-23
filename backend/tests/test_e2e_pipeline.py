import os
import sys
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

os.environ.setdefault("BOMATIC_API_KEY", "test-e2e-key")

import openpyxl
import pytest
from fastapi.testclient import TestClient
from jose import jwt
from docx import Document as DocxDocument

from app.config import get_settings
from app.main import app


pytestmark = pytest.mark.e2e

FIXTURE_DIR = Path(__file__).parent / "fixtures"
SAMPLE_RFP_PDF = FIXTURE_DIR / "sample_rfp.pdf"
SAMPLE_BOQ_XLSX = FIXTURE_DIR / "sample_boq.xlsx"
OUTPUT_DIRS = {
    "e2": Path(__file__).resolve().parents[1] / "app/engines/e2/output",
    "e3": Path(__file__).resolve().parents[1] / "app/engines/e3/output",
    "e4": Path(__file__).resolve().parents[1] / "app/engines/e4/output",
    "e5": Path(__file__).resolve().parents[1] / "app/engines/e5/output",
}


def _api_key() -> str:
    return os.environ.get("BOMATIC_API_KEY", "test-e2e-key")


def _auth_headers(client: TestClient, email_prefix: str) -> dict[str, str]:
    settings = get_settings()
    email = f"{email_prefix}-{uuid.uuid4().hex[:8]}@example.com"
    register = client.post(
        "/api/auth/register",
        json={
            "email": email,
            "password": "e2e-test-password",
            "full_name": "E2E Test User",
        },
    )
    assert register.status_code == 201, register.text

    me = client.get(
        "/api/auth/me",
        headers={
            "Authorization": f"Bearer {register.json()['access_token']}",
            "X-API-Key": _api_key(),
        },
    )
    assert me.status_code == 200, me.text
    user_id = me.json()["id"]
    owner_token = jwt.encode(
        {
            "sub": user_id,
            "exp": datetime.now(timezone.utc) + timedelta(minutes=15),
        },
        settings.jwt_secret_key,
        algorithm=settings.jwt_algorithm,
    )
    return {
        "Authorization": f"Bearer {owner_token}",
        "X-API-Key": _api_key(),
    }


def _create_opportunity(client: TestClient, *, mode: str, email_prefix: str, filename: str, content: bytes) -> tuple[str, dict[str, str]]:
    return _create_opportunity_with_files(
        client,
        mode=mode,
        email_prefix=email_prefix,
        files=[(filename, content)],
    )


def _create_opportunity_with_files(
    client: TestClient,
    *,
    mode: str,
    email_prefix: str,
    files: list[tuple[str, bytes]],
) -> tuple[str, dict[str, str]]:
    headers = _auth_headers(client, email_prefix)
    opportunity_id = f"E2E-{uuid.uuid4().hex[:10].upper()}"
    file_tuples = []
    for filename, content in files:
        mime_type = (
            "application/pdf"
            if filename.endswith(".pdf")
            else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            if filename.endswith(".docx")
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_tuples.append(("files", (filename, BytesIO(content), mime_type)))
    response = client.post(
        "/api/v1/rfp/packages",
        data={
            "opportunity_id": opportunity_id,
            "project_name": f"E2E {mode.upper()} Project",
            "mode": mode,
        },
        files=file_tuples,
        headers=headers,
    )
    assert response.status_code == 201, response.text
    return opportunity_id, headers


def _context_docx_bytes(text: str) -> bytes:
    document = DocxDocument()
    for paragraph in text.splitlines():
        document.add_paragraph(paragraph)
    buffer = BytesIO()
    document.save(buffer)
    return buffer.getvalue()


def _docx_contains_placeholder(path: Path) -> bool:
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if not name.startswith("word/") or not name.endswith(".xml"):
                continue
            if "[Section content to be completed manually]" in archive.read(name).decode("utf-8", errors="ignore"):
                return True
    return False


def _priced_rows(path: Path) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        rows: list[tuple] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                if len(row) < 5:
                    continue
                unit_price = row[3]
                line_total = row[4]
                if isinstance(unit_price, (int, float)) and isinstance(line_total, (int, float)):
                    rows.append(row)
        return rows
    finally:
        workbook.close()


def _write_response_workbook(path: Path, questions: list[dict]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "RFI Questionnaire"
    worksheet.append(["RFI ID", "Category", "Priority", "Question", "Rationale", "Answer Type", "Response"])
    for question in questions:
        answer_type = str(question.get("expected_answer_type", "text")).lower()
        if answer_type == "number":
            answer = "250"
        elif answer_type == "yes_no":
            answer = "Yes"
        elif answer_type == "attachment":
            answer = "Attached network diagram"
        else:
            answer = f"Response for {question.get('question', 'question')}"
        worksheet.append([
            question.get("id", ""),
            question.get("category", "General"),
            question.get("priority", "must_have"),
            question.get("question", ""),
            question.get("rationale", ""),
            question.get("expected_answer_type", "text"),
            answer,
        ])
    workbook.save(path)


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as test_client:
        yield test_client


def test_e2e_rfp_pipeline_generates_docx_and_priced_xlsx(client, monkeypatch):
    assert SAMPLE_RFP_PDF.exists(), "Missing sample_rfp.pdf fixture"
    assert SAMPLE_BOQ_XLSX.exists(), "Missing sample_boq.xlsx fixture"

    opportunity_id, headers = _create_opportunity_with_files(
        client,
        mode="rfp",
        email_prefix="e2e-rfp",
        files=[
            ("sample_rfp.pdf", SAMPLE_RFP_PDF.read_bytes()),
            ("sample_rfp_context.docx", _context_docx_bytes((FIXTURE_DIR / "sample_rfp_test.txt").read_text(encoding="utf-8"))),
        ],
    )

    e1_run = client.post(f"/api/e1/{opportunity_id}/run", headers=headers)
    assert e1_run.status_code == 200, e1_run.text

    e2_run = client.post(
        "/api/e2/analyze",
        data={
            "rfp_session_id": opportunity_id,
            "pasted_text": "2, Cisco Catalyst 9300 switch\n1, Cisco ASA 5516-X firewall",
            "target_currency": "SAR",
            "vat_country": "SA",
            "vendor_discount_pct": "0.30",
            "inhouse_margin_pct": "0.10",
            "selling_mode": "margin",
            "selling_pct": "0.25",
        },
        files={
            "boq_template": (
                "sample_boq.xlsx",
                SAMPLE_BOQ_XLSX.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )
    assert e2_run.status_code == 200, e2_run.text
    e2_payload = e2_run.json()
    e2_output = OUTPUT_DIRS["e2"] / Path(e2_payload["output_file"]).name
    assert e2_output.exists(), str(e2_output)
    assert _priced_rows(e2_output), "E2 output workbook did not contain any priced rows."

    def _fake_generate_narratives(e1_data, e2_data, sections, gbb_tier="better"):
        content_by_title = {
            "Cover Page": "Cover page prepared for submission.",
            "Cover Letter / Introduction": "Cover letter introducing the proposal and delivery team.",
            "Executive Summary": "Executive summary for the proposal. This proposal addresses the client requirements with a clear delivery model.",
            "Understanding of Customer Requirements": "Requirement analysis summary. The proposal covers the mandatory and conditional requirements identified in the RFP.",
            "Proposed Solution": "Proposed solution narrative. The solution uses the matched BoM and design components to satisfy the stated requirements.",
            "Technical Specifications": "Detailed technical specifications for the proposed solution.",
            "Implementation Approach": "Implementation approach and delivery sequencing for the project.",
            "Service Levels & Support": "Service levels and support arrangements for the engagement.",
            "Commercial Proposal": "Commercial proposal with pricing summary derived from the BoM.",
            "Scope, Assumptions, Exclusions, Dependencies": "Scope, assumptions, exclusions, and dependencies for the proposal.",
            "Compliance Matrix": "Compliance matrix summary aligned to the extracted requirements.",
            "References / Case Studies": "Selected references and case studies demonstrating similar deployments.",
            "Company Profile": "Company profile and delivery credentials.",
            "Appendices": "Appendices containing supporting material for the submission.",
            "Signature Page": "Signature page for authorised sign-off.",
        }
        return {section.id: content_by_title.get(section.title, f"{section.title} content.") for section in sections}

    monkeypatch.setattr("app.engines.e3.pipeline.generate_narratives", _fake_generate_narratives)

    e3_run = client.post(
        "/api/e3/generate",
        data={
            "rfp_session_id": opportunity_id,
            "gbb_tier": "better",
            "allow_placeholders": "false",
        },
        headers=headers,
    )
    assert e3_run.status_code == 200, e3_run.text
    e3_payload = e3_run.json()
    e3_output = OUTPUT_DIRS["e3"] / Path(e3_payload["output_file"]).name
    assert e3_output.exists(), str(e3_output)
    assert not _docx_contains_placeholder(e3_output), "E3 output DOCX still contains placeholder tokens."


def test_e2e_quick_bom_pastes_line_items_and_prices_workbook(client):
    assert SAMPLE_BOQ_XLSX.exists(), "Missing sample_boq.xlsx fixture"

    pasted_lines = "\n".join(
        [
            "2, Cisco Catalyst 9300 switch",
            "1, Cisco ASA 5516-X firewall",
            "5, Wireless access point",
        ]
    )

    response = client.post(
        "/api/e2/analyze",
        data={
            "rfp_session_id": "",
            "pasted_text": pasted_lines,
            "target_currency": "SAR",
            "vat_country": "SA",
            "vendor_discount_pct": "0.30",
            "inhouse_margin_pct": "0.10",
            "selling_mode": "margin",
            "selling_pct": "0.25",
        },
        files={
            "boq_template": (
                "sample_boq.xlsx",
                SAMPLE_BOQ_XLSX.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-API-Key": _api_key()},
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    output_path = OUTPUT_DIRS["e2"] / Path(payload["output_file"]).name
    assert output_path.exists(), str(output_path)

    priced_rows = _priced_rows(output_path)
    assert priced_rows, "Quick BoM workbook has no priced rows."
    assert any("Cisco" in str(row[1]) for row in priced_rows if len(row) > 1)


def test_e2e_rfi_pipeline_generates_questionnaire_response_and_components(client):
    assert SAMPLE_RFP_PDF.exists(), "Missing sample_rfp.pdf fixture"

    opportunity_id, headers = _create_opportunity(
        client,
        mode="rfi",
        email_prefix="e2e-rfi",
        filename="sample_rfp.pdf",
        content=SAMPLE_RFP_PDF.read_bytes(),
    )

    e4_generate = client.post(
        "/api/e4/generate",
        data={"rfp_session_id": opportunity_id, "project_name": "RFI E2E Project"},
        headers={"X-API-Key": _api_key()},
    )
    assert e4_generate.status_code == 200, e4_generate.text
    e4_payload = e4_generate.json()
    questions = e4_payload.get("questions", [])
    assert questions, "E4 questionnaire did not generate any questions."

    response_path = Path(__file__).parent / f"rfi_response_{uuid.uuid4().hex}.xlsx"
    try:
        _write_response_workbook(response_path, questions)
        upload = client.post(
            f"/api/e4/{opportunity_id}/responses",
            files={
                "response_file": (
                    response_path.name,
                    response_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-API-Key": _api_key()},
        )
        assert upload.status_code == 200, upload.text

        e5_generate = client.post(
            "/api/e5/generate",
            data={"rfp_session_id": opportunity_id},
            headers={"X-API-Key": _api_key()},
        )
        assert e5_generate.status_code == 200, e5_generate.text
        e5_payload = e5_generate.json()
        component_artifact = e5_payload.get("component_artifact", {})
        components = component_artifact.get("components", [])
        assert components, "E5 component list is empty."
    finally:
        if response_path.exists():
            response_path.unlink()
